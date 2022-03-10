import argparse
import csv
import io
import logging
import os
import random
import sys
import time
import warnings
from logging.handlers import RotatingFileHandler

import chromedriver_autoinstaller
import openpyxl
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By

LOG_FILE = 'log.log'
DEFAULT_HEADLESS = False
MAX_QUERIES = 100


def create_webdriver(headless=False):
    chromedriver_autoinstaller.install()
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-ssl-errors')
    options.add_argument('--ignore-certificate-errors-spki-list')
    options.add_argument('log-level=2')
    if headless:
        options.add_argument('--headless')
    return webdriver.Chrome(options=options)


def create_logger(log_file):
    log_format = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
    logger = logging.getLogger('google_reviews_scraper')
    logger.setLevel(logging.DEBUG)

    # log to file
    log_file_handler = RotatingFileHandler(log_file, mode='a', maxBytes=1000000, backupCount=1, encoding='utf-8')
    log_file_handler.setLevel(logging.DEBUG)
    log_file_handler.setFormatter(log_format)
    logger.addHandler(log_file_handler)

    # log to console
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(log_format)
    logger.addHandler(console_handler)
    return logger


def open_workbook_in_memory(input_file_path):
    """ Return an opened workbook.

        Openpyxl sometimes doesn't close the file properly and it remains blocked, so it's copied to memory first
    """
    with open(input_file_path, "rb") as f:
        in_mem_file = io.BytesIO(f.read())

    workbook = openpyxl.load_workbook(in_mem_file, read_only=True, data_only=True)
    return workbook


def parse_queries_from_xlsx_file(input_file_path):
    """ Return the non-empty cell values from the first column in each sheet. """
    queries = list()
    warnings.simplefilter("ignore")  # Suppress openpyxl UserWarnings
    workbook = open_workbook_in_memory(input_file_path)
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        sheet.reset_dimensions()
        for row in sheet.iter_rows():
            if row and row[0] is not None and row[0].value:  # the first cell in the row is not empty
                queries.append(row[0].value)
    if queries and queries[0] == 'Agent Name':  # remove header, if present
        queries = queries[1:]
    return queries


def search_query_on_google(driver, query):
    search_field = driver.find_element(By.XPATH, '//input[@name="q"]')
    search_field.clear()
    search_field.send_keys(query)
    time.sleep(random.uniform(0.2, 2))
    search_field.submit()


def parse_star_score(driver):
    el = driver.find_element(By.XPATH, '//div[contains(@data-attrid, "star_score")]//span')
    return el.text.replace(',', '.')


def parse_number_of_reviews(driver):
    el = driver.find_element(By.XPATH, '//a[contains(@data-sort_by, "qualityScore")]')
    return el.text.split(' ')[0]


def parse_search_results(driver, query, logger):
    search_results = dict()
    search_results['query'] = query
    try:
        search_query_on_google(driver=driver, query=query)
        search_results['url'] = driver.current_url
    except Exception as e:
        logger.exception(e)
        return search_results

    try:
        search_results['star_score'] = parse_star_score(driver)
        search_results['reviews_count'] = parse_number_of_reviews(driver)
    except NoSuchElementException:
        logger.debug(f'No review elements found for query {query}')
    except Exception as e:
        logger.exception(e)
    return search_results


def main():
    parser = argparse.ArgumentParser(
        description='Google reviews scraper, searching queries from the provided .xlsx file '
                    'and storing rating and the number of reviews from the side panel',
        usage='python review_parser.py input_file.xlsx [-o output_file.xlsx]'
    )
    parser.add_argument('input_file', help='path to the input file')
    parser.add_argument('-o', dest='output_file', default='output.csv', help='[optional] path to the output file.')
    options = parser.parse_args()

    if not os.path.isfile(options.input_file):
        raise FileNotFoundError('Couldn\'t find input file at the provided path')

    logger = create_logger(log_file=LOG_FILE)
    try:
        driver = create_webdriver(headless=DEFAULT_HEADLESS)
        queries = parse_queries_from_xlsx_file(input_file_path=options.input_file)
        queries = queries[:MAX_QUERIES]
    except Exception as e:
        logger.exception(e)
        sys.exit(1)

    logger.info(f'Parsed {len(queries)} from the input file')

    try:
        with open(options.output_file, 'w', newline='', encoding='utf-8') as outfile:
            seen_search_results = dict()
            fieldnames = ['query', 'star_score', 'reviews_count', 'url']
            writer = csv.DictWriter(outfile, fieldnames=fieldnames, dialect='excel', delimiter=';')
            writer.writeheader()

            driver.get('https://google.com')
            for query in queries:
                if query in seen_search_results:
                    search_results = seen_search_results[query]
                    logger.debug(f'Using cached results: {search_results}')
                else:
                    time.sleep(random.uniform(3, 5))
                    search_results = parse_search_results(driver=driver, query=query, logger=logger)
                    logger.debug(f'Parsed query results: {search_results}')
                    seen_search_results[query] = search_results
                writer.writerow(search_results)
    except Exception as e:
        logger.exception(e)
        sys.exit(1)


if __name__ == '__main__':
    main()
